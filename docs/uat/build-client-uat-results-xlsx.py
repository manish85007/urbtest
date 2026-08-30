#!/usr/bin/env python3
"""Build credential-free Client UAT Excel workbook for individual testers."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent / "Urb-TecTrack-Client-UAT-Results.xlsx"

GREEN = "0F5C3A"
GREEN_FILL = PatternFill("solid", fgColor=GREEN)
HEADER_FILL = PatternFill("solid", fgColor="E8F2EC")
LIGHT_FILL = PatternFill("solid", fgColor="F7FAF8")
WARN_FILL = PatternFill("solid", fgColor="FFF8E7")
WHITE = "FFFFFF"
THIN = Border(
    left=Side(style="thin", color="C5D5CC"),
    right=Side(style="thin", color="C5D5CC"),
    top=Side(style="thin", color="C5D5CC"),
    bottom=Side(style="thin", color="C5D5CC"),
)

RESULT_OPTIONS = '"Pass,Fail,N/A,Blocked"'
SEVERITY_OPTIONS = '"Blocker,Major,Minor,Cosmetic"'

CASES: list[tuple[str, str, str, str]] = [
    # section, id, step, expected
    ("C0 Sign-in, policies, navigation", "C0.1", "Open the portal URL. Sign in with the account issued to you separately.", "Login succeeds."),
    ("C0 Sign-in, policies, navigation", "C0.2", "If Accept policies appears: open Terms and Privacy, then accept.", "Gate clears. Home welcomes you by name."),
    ("C0 Sign-in, policies, navigation", "C0.3", "Review main navigation.", "Client areas only: Home, My Requests, Recycling Heroes, Sustainability, Reports. No Masters, Audit, Capacity, Compliance."),
    ("C0 Sign-in, policies, navigation", "C0.4", "Open Your profile (avatar / name).", "Role is Client User. Organisation matches yours. Change password available. Two-factor is NOT shown for clients."),
    ("C0 Sign-in, policies, navigation", "C0.5", "Sign out, then sign in again.", "Returns to Home. Policies not re-prompted unless Urbeno published a new version."),
    ("C1 Data isolation (critical)", "C1.1", "Open My Requests. Scan the list.", "Only your organisation's requests. No other company's request IDs."),
    ("C1 Data isolation (critical)", "C1.2", "If Urbeno supplies another organisation's request ID, open it via the address bar.", "Access refused or redirected. Must not show another company's details."),
    ("C1 Data isolation (critical)", "C1.3", "Open Reports (Request Summary / Invoice Register).", "Rows are your organisation only. No MRN Register for clients."),
    ("C1 Data isolation (critical)", "C1.4", "Try /masters, /audit, /compliance, /capacity on the portal URL.", "Redirected or access denied. No restricted data shown."),
    ("C2 Raise a collection request (Stage 1)", "C2.1", "Home or My Requests → + New Request.", "Form opens. Organisation implied. Site / location required."),
    ("C2 Raise a collection request (Stage 1)", "C2.2", "Submit with required fields empty.", "Form does not submit; clear validation messages appear."),
    ("C2 Raise a collection request (Stage 1)", "C2.3", "Fill pickup location, approx. quantity & weight, at least one line item; optional PO/notes. Submit.", "New REQ-##### at Stage 1 (awaiting acknowledgement). Record ID in Session sheet."),
    ("C2 Raise a collection request (Stage 1)", "C2.4", "Check header / actions on the new request.", "Organisation, site, raised date, raised-by correct. No Acknowledge / Assign Vehicle / Weigh / Raise Invoice."),
    ("C3 Changes requested & resubmit (needs Urbeno; else N/A)", "C3.1", "After Urbeno requests changes: open the request.", "Status / badge shows changes requested; admin note visible."),
    ("C3 Changes requested & resubmit (needs Urbeno; else N/A)", "C3.2", "Edit pickup details → save / resubmit.", "Sent back to Urbeno. Stage remains awaiting acknowledgement. Still cannot Acknowledge."),
    ("C4 Visibility while staff process (Stages 2–8)", "C4.1", "After ack / vehicle / weighment / invoice: refresh the request.", "Stage badge advances. Vehicle and invoice details appear when ready."),
    ("C4 Visibility while staff process (Stages 2–8)", "C4.2", "Inspect invoice / processing panels.", "Progress visible. No MRN number, Create MRN, or factory security-officer MRN fields."),
    ("C4 Visibility while staff process (Stages 2–8)", "C4.3", "After recycling / Form 6 (factory).", "Processing progressed; you are not asked to issue Form 6."),
    ("C4 Visibility while staff process (Stages 2–8)", "C4.4", "After certificate upload (Urbeno).", "Certificate reference/date visible. No Upload Certificate for you."),
    ("C4 Visibility while staff process (Stages 2–8)", "C4.5", "Try Review & Close before payment is complete (if button appears).", "Close unavailable or refused until payment is recorded."),
    ("C5 Close the request (Stage 9)", "C5.1", "When certificate exists AND payment is recorded: Review & Close.", "Closure confirmation modal opens."),
    ("C5 Close the request (Stage 9)", "C5.2", "Acknowledge closure.", "Invoice/request closed. Stage shows Closed / Stage 9."),
    ("C5 Close the request (Stage 9)", "C5.3", "Check Home completed / closed counts.", "Counts reflect the closed work."),
    ("C6 Home, Sustainability, Recycling Heroes", "C6.1", "Open Home.", "Open/completed tiles and impact figures load. + New Request available."),
    ("C6 Home, Sustainability, Recycling Heroes", "C6.2", "Open Sustainability.", "Impact figures load; explanatory text / Impact PDF (if offered) open without error. Reflects closed work."),
    ("C6 Home, Sustainability, Recycling Heroes", "C6.3", "Open Recycling Heroes.", "Tree / tonnage view loads."),
    ("C7 Reports & export", "C7.1", "Reports → Request Summary → Export CSV (if available).", "File downloads; contains only your organisation's rows."),
    ("C7 Reports & export", "C7.2", "Try Invoice Register, Certificate Log, Sustainability, Recycling Heroes reports.", "Each runs without error. No MRN register."),
    ("C7 Reports & export", "C7.3", "Change period (FY / month) if shown.", "Figures refresh; still no other organisation's data."),
    ("C8 Profile / password (optional — do last)", "C8.1", "Profile → try a weak password (e.g. short).", "Refused; policy requires 10+ chars, upper, lower, digit."),
    ("C8 Profile / password (optional — do last)", "C8.2", "Successful password change.", "Only if Urbeno asked you to change it; otherwise mark N/A."),
]


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = GREEN_FILL
        cell.font = Font(color=WHITE, bold=True, name="Calibri", size=11)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN


def autofit(ws, widths: dict[int, int]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def add_title(ws, text: str, row=1, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row, 1, text)
    cell.font = Font(name="Calibri", size=16, bold=True, color=GREEN)
    cell.alignment = Alignment(vertical="center")


def build():
    wb = Workbook()

    # ——— Instructions ———
    ws = wb.active
    ws.title = "Instructions"
    add_title(ws, "Urb TecTrack — Client UAT Results (Tester workbook)", cols=2)
    ws.row_dimensions[1].height = 28

    instructions = [
        "",
        "Purpose",
        "Capture Pass / Fail results for the client (requestor) portal during User Acceptance Testing.",
        "",
        "Important — credentials",
        "This workbook does NOT contain usernames or passwords.",
        "Urbeno issues sign-in details separately. Do not write passwords in this file.",
        "",
        "How to use",
        "1. Complete the Session sheet with your organisation and tester details.",
        "2. Review Test Parameters for expected behaviour.",
        "3. Work through Checklist — select Result from the dropdown for every case.",
        "4. Log every Fail or Blocked item on the Findings sheet.",
        "5. Complete Summary & Sign-off, then email this file to info@urbeno.in.",
        "",
        "Result meanings",
        "Pass — behaviour matches Expected",
        "Fail — wrong or missing vs Expected (must add a Findings row)",
        "N/A — cannot run in this session (state why in Notes)",
        "Blocked — blocked by another Fail or environment issue",
        "",
        "Severity (Findings)",
        "Blocker — cannot raise/view/close own requests; sees other org data; security concern",
        "Major — business rule broken; workaround exists",
        "Minor — awkward UI; process still completable",
        "Cosmetic — spelling / layout only",
        "",
        "Portal",
        "https://uat.urbeno.in",
        "",
        "Return",
        "Email: info@urbeno.in",
        "Subject: Urb TecTrack UAT — Client findings — [Your Organisation] — [Date]",
        "Companion visual guide (no credentials): Urb-TecTrack-Client-Access-Visual-Guide.pdf",
    ]
    for i, line in enumerate(instructions, start=2):
        cell = ws.cell(i, 1, line)
        cell.font = Font(name="Calibri", size=11, bold=line in {
            "Purpose", "Important — credentials", "How to use", "Result meanings",
            "Severity (Findings)", "Portal", "Return",
        } or line.startswith("Urb TecTrack"))
        if line.startswith("This workbook does NOT") or line.startswith("Urbeno issues"):
            cell.fill = WARN_FILL
        cell.alignment = Alignment(wrap_text=True)
    autofit(ws, {1: 100, 2: 20})
    ws.print_title_rows = "1:1"

    # ——— Session ———
    ws = wb.create_sheet("Session")
    add_title(ws, "Session details", cols=2)
    ws.row_dimensions[1].height = 26
    ws.cell(2, 1, "Credentials are issued separately and must not be stored in this workbook.").fill = WARN_FILL
    ws.merge_cells("A2:B2")

    session_rows = [
        ("Organisation", ""),
        ("Tester name", ""),
        ("Tester email", ""),
        ("Date(s) of testing (IST)", ""),
        ("Browser + version", ""),
        ("Device", ""),
        ("Portal URL", "https://uat.urbeno.in"),
        ("Account used (email only — issued separately)", ""),
        ("Build / release note (if provided)", ""),
        ("Request ID created during UAT (REQ-…)", ""),
        ("Other-org request ID for tenancy test (if supplied by Urbeno)", ""),
    ]
    ws.cell(4, 1, "Field").font = Font(bold=True, color=WHITE)
    ws.cell(4, 2, "Your entry").font = Font(bold=True, color=WHITE)
    style_header_row(ws, 4, 2)
    for i, (field, default) in enumerate(session_rows, start=5):
        ws.cell(i, 1, field).border = THIN
        ws.cell(i, 1).fill = HEADER_FILL
        c = ws.cell(i, 2, default)
        c.border = THIN
        c.fill = LIGHT_FILL
    autofit(ws, {1: 55, 2: 55})

    notes = [
        "",
        "Before you start",
        "• Use only the account Urbeno issued to you.",
        "• Prefer a private/incognito window if you also use other portal roles.",
        "• On first sign-in you may need to accept Terms and Privacy.",
        "• Do not change a shared password unless Urbeno asks you to.",
    ]
    for i, line in enumerate(notes, start=17):
        ws.cell(i, 1, line).font = Font(name="Calibri", size=11, bold=line == "Before you start")

    # ——— Test Parameters ———
    ws = wb.create_sheet("Test Parameters")
    add_title(ws, "Test parameters (reference — no credentials)", cols=2)
    style_header_row(ws, 3, 2)
    ws.cell(3, 1, "Parameter")
    ws.cell(3, 2, "Value / guidance")
    style_header_row(ws, 3, 2)
    params = [
        ("Portal URL", "https://uat.urbeno.in"),
        ("Role under test", "Client User (requestor)"),
        ("Expected navigation", "Home, My Requests, Recycling Heroes, Sustainability, Reports"),
        ("Must NOT appear in navigation", "Masters, Audit, Capacity, Compliance"),
        ("Sample new-request data", "Pickup location e.g. Loading bay; Approx. qty e.g. 12; Approx. weight e.g. 75 kg; at least one line item"),
        ("Tenancy rule", "Only your organisation's requests and reports"),
        ("Staff-only actions (must not appear for you)", "Acknowledge, Assign Vehicle, Weigh, Raise Invoice, Create MRN, Upload Certificate, Form 6 issue"),
        ("Close rule", "Review & Close only after certificate AND payment are complete"),
        ("Password policy (if changing password)", "10+ characters; upper-case; lower-case; digit; cannot repeat last 5 passwords"),
        ("Two-factor authentication", "Not used for client users in this portal"),
        ("Return findings to", "info@urbeno.in"),
        ("Email subject template", "Urb TecTrack UAT — Client findings — [Organisation] — [Date]"),
    ]
    for i, (k, v) in enumerate(params, start=4):
        ws.cell(i, 1, k).border = THIN
        ws.cell(i, 1).fill = HEADER_FILL
        c = ws.cell(i, 2, v)
        c.border = THIN
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 32
    autofit(ws, {1: 42, 2: 75})

    # ——— Checklist ———
    ws = wb.create_sheet("Checklist")
    add_title(ws, "Client portal checklist — mark Result for every case", cols=7)
    ws.row_dimensions[1].height = 26
    headers = ["Section", "Case ID", "Step", "Expected", "Result", "Initials", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h)
    style_header_row(ws, 3, 7)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{3 + len(CASES)}"

    result_dv = DataValidation(type="list", formula1=RESULT_OPTIONS, allow_blank=True)
    result_dv.error = "Select Pass, Fail, N/A, or Blocked"
    result_dv.errorTitle = "Invalid result"
    ws.add_data_validation(result_dv)

    for i, (section, case_id, step, expected) in enumerate(CASES, start=4):
        ws.cell(i, 1, section).border = THIN
        ws.cell(i, 2, case_id).border = THIN
        ws.cell(i, 2).font = Font(bold=True, name="Calibri")
        for col, val in ((3, step), (4, expected)):
            cell = ws.cell(i, col, val)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        res = ws.cell(i, 5, "")
        res.border = THIN
        res.fill = LIGHT_FILL
        result_dv.add(res)
        for col in (6, 7):
            cell = ws.cell(i, col, "")
            cell.border = THIN
            cell.fill = LIGHT_FILL
        if i % 2 == 0:
            for col in range(1, 5):
                if not ws.cell(i, col).fill or ws.cell(i, col).fill.fgColor is None:
                    ws.cell(i, col).fill = PatternFill("solid", fgColor="FBFCFB")
        ws.row_dimensions[i].height = 48

    last = 3 + len(CASES)
    autofit(ws, {1: 42, 2: 10, 3: 48, 4: 48, 5: 12, 6: 10, 7: 28})

    # Counts helpers
    ws.cell(last + 2, 1, "Quick counts (auto)").font = Font(bold=True, color=GREEN)
    ws.cell(last + 3, 1, "Pass")
    ws.cell(last + 3, 2, f'=COUNTIF(E4:E{last},"Pass")')
    ws.cell(last + 4, 1, "Fail")
    ws.cell(last + 4, 2, f'=COUNTIF(E4:E{last},"Fail")')
    ws.cell(last + 5, 1, "N/A")
    ws.cell(last + 5, 2, f'=COUNTIF(E4:E{last},"N/A")')
    ws.cell(last + 6, 1, "Blocked")
    ws.cell(last + 6, 2, f'=COUNTIF(E4:E{last},"Blocked")')
    ws.cell(last + 7, 1, "Blank (not yet tested)")
    ws.cell(last + 7, 2, f'=COUNTBLANK(E4:E{last})')

    # ——— Findings ———
    ws = wb.create_sheet("Findings")
    add_title(ws, "Findings log — one row per Fail / Blocked / notable observation", cols=8)
    headers = [
        "Finding ID",
        "Case ID",
        "Severity",
        "Summary (what you saw vs expected)",
        "Screenshot / attachment name",
        "Steps to reproduce",
        "Status",
        "Owner / notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h)
    style_header_row(ws, 3, 8)
    ws.freeze_panes = "A4"

    sev_dv = DataValidation(type="list", formula1=SEVERITY_OPTIONS, allow_blank=True)
    status_dv = DataValidation(type="list", formula1='"Open,Fixed,Waived,Deferred"', allow_blank=True)
    ws.add_data_validation(sev_dv)
    ws.add_data_validation(status_dv)

    for i in range(4, 24):
        ws.cell(i, 1, f"F-{i - 3:03d}").border = THIN
        for c in range(2, 9):
            cell = ws.cell(i, c, "")
            cell.border = THIN
            cell.fill = LIGHT_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sev_dv.add(ws.cell(i, 3))
        status_dv.add(ws.cell(i, 7))
        ws.cell(i, 7, "Open")
        ws.row_dimensions[i].height = 36

    autofit(ws, {1: 12, 2: 10, 3: 12, 4: 45, 5: 28, 6: 35, 7: 12, 8: 22})
    ws.cell(25, 1, "Quote exact on-screen error or success messages in the Summary column.").fill = WARN_FILL
    ws.merge_cells("A25:H25")

    # ——— Summary & Sign-off ———
    ws = wb.create_sheet("Summary & Sign-off")
    add_title(ws, "Summary and sign-off", cols=2)
    ws.cell(3, 1, "Metric")
    ws.cell(3, 2, "Count / answer")
    style_header_row(ws, 3, 2)

    last_case_row = 3 + len(CASES)
    summary = [
        ("Cases on Checklist", str(len(CASES))),
        ("Pass", f'=COUNTIF(Checklist!E4:E{last_case_row},"Pass")'),
        ("Fail", f'=COUNTIF(Checklist!E4:E{last_case_row},"Fail")'),
        ("N/A", f'=COUNTIF(Checklist!E4:E{last_case_row},"N/A")'),
        ("Blocked", f'=COUNTIF(Checklist!E4:E{last_case_row},"Blocked")'),
        ("Not yet tested", f'=COUNTBLANK(Checklist!E4:E{last_case_row})'),
        ("Open Blockers (enter manually)", ""),
        ("Open Majors (enter manually)", ""),
        ("Blockers found? (No / Yes — IDs)", ""),
        ("Fit for production as a client user? (Yes / Yes with waivers / No)", ""),
        ("Overall comments", ""),
    ]
    for i, (k, v) in enumerate(summary, start=4):
        ws.cell(i, 1, k).border = THIN
        ws.cell(i, 1).fill = HEADER_FILL
        c = ws.cell(i, 2, v)
        c.border = THIN
        c.fill = LIGHT_FILL
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 22

    fit_dv = DataValidation(
        type="list",
        formula1='"Yes,Yes with waivers,No"',
        allow_blank=True,
    )
    ws.add_data_validation(fit_dv)
    fit_dv.add(ws.cell(13, 2))

    ws.cell(16, 1, "Declaration").font = Font(bold=True, color=GREEN)
    ws.cell(
        17,
        1,
        "I confirm I executed this checklist on the UAT portal and that the results and findings are accurate. "
        "I have not recorded any passwords in this workbook.",
    )
    ws.merge_cells("A17:B17")
    ws.row_dimensions[17].height = 40
    ws.cell(17, 1).alignment = Alignment(wrap_text=True)

    sign = [
        ("Tester name", ""),
        ("Organisation", ""),
        ("Signature / typed name", ""),
        ("Date", ""),
        ("Client sponsor (optional)", ""),
    ]
    ws.cell(19, 1, "Field")
    ws.cell(19, 2, "Entry")
    style_header_row(ws, 19, 2)
    for i, (k, v) in enumerate(sign, start=20):
        ws.cell(i, 1, k).border = THIN
        ws.cell(i, 1).fill = HEADER_FILL
        c = ws.cell(i, 2, v)
        c.border = THIN
        c.fill = LIGHT_FILL

    ws.cell(26, 1, "Return this file to info@urbeno.in").fill = WARN_FILL
    ws.merge_cells("A26:B26")
    autofit(ws, {1: 60, 2: 45})

    wb.save(OUT)
    print(f"Wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    build()
